from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam (1).xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

count = 0

for row in range(2, max_row + 1):
    address = ws['D' + str(row)].value
    number = ws['L' + str(row)].value

    if isinstance(address, str) and address.startswith('Ain') and isinstance(number, (int, float)) and number < 40:
        count += 1

print(count)



