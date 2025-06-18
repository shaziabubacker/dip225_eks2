from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam (1).xlsx', data_only=True)
ws = wb['Lapa_0']
max_row = ws.max_row

sum_total = 0

for row in range(2, max_row + 1):
    client_type = ws['F' + str(row)].value      # Client column
    qty = ws['L' + str(row)].value              # Quantity
    total_amount = ws['N' + str(row)].value     # Total price

    # Check conditions
    if client_type == 'Korporatīvais' and qty is not None and total_amount is not None:
        try:
            qty = float(qty)
            if 40 <= qty <= 50:
                sum_total += float(total_amount)
