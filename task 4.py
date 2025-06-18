from openpyxl import load_workbook

# Load workbook and worksheet
wb = load_workbook('sagatave_eksamenam (1).xlsx')
ws = wb['Lapa_0']

sum_prices = 0
num_items = 0
rows = ws.max_row

for i in range(2, rows + 1):
    product_name = ws['I' + str(i)].value    # Product name
    product_price = ws['K' + str(i)].value   # Price

    if product_name and 'LaserJet' in product_name and isinstance(product_price, (int, float)):
        sum_prices += product_price
        num_items += 1

if num_items > 0:
    average_price = round(sum_prices / num_items)
    print(average_price)
else:
    print(0)
