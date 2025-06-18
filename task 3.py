from openpyxl import load_workbook

# Load workbook (use correct file name with (1))
wb = load_workbook('sagatave_eksamenam (1).xlsx')

# Select the correct sheet
ws = wb['Lapa_0']

# Initialize counter
count = 0

# Loop through all rows starting from row 2
for row in range(2, ws.max_row + 1):
    # Read values from columns D and E
    address = ws[f'D{row}'].value
    city = ws[f'E{row}'].value

    # Check if address contains 'Adulienas iela' and city matches
    if isinstance(address, str) and 'Adulienas iela' in address and city in ('Valmiera', 'Saulkrasti'):
        count += 1

# Print final count
print(f"Total matching rows: {count}")
