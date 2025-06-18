from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam (1).xlsx')
ws = wb['Lapa_0']
max_row = ws.max_row

count = 0

for row in range(2, max_row + 1):
    priority = ws['H' + str(row)].value       # Prioritāte
    delivery_date = ws['J' + str(row)].value  # Piegādes datums

    try:
        if priority == 'High' and delivery_date.year == 2015:
            count += 1
    except:
        # If delivery_date is not a date (e.g. None or string), skip it
        continue

print(count)
